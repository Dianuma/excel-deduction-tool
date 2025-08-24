import os
import warnings
from openpyxl import load_workbook
import xlrd
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.alert import Alert
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.filedialog import askopenfilename
# 공제 불공제 변경 시스템
class hometax_program():
    def __init__(self):
        # 속성 초기화
        self.driver = self.open_chrome()
        self.ID_password = None

    def change_deduction(self, data, master):
        # 공제 불공제 변경
        count_page=1
        total_deduction=0
        element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/thead/tr/th[1]/input')
        self.driver.execute_script("arguments[0].click();", element)
        for i in range(2,len(data)):
            table_order=str((i-2)%20+1)
            day=self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[2]').text
            franchise_id=self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[5]').text
            name=self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[6]').text
            total=self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[10]').text
            if day==data[i][0] and franchise_id==data[i][1] and name==data[i][2] and int(total.replace(',',''))==int(data[i][3]):
                try:
                    if data[i][4].replace(' ','')=='공제':
                        Select(self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[14]/div/div/select')).select_by_index(0)
                    else:
                        Select(self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/tbody/tr['+table_order+']/td[14]/div/div/select')).select_by_index(1)
                    progress=((i-2)/(len(data)-2))*100
                    self.display_progress(str(i - 1)+" - "+data[i][2]+" : "+data[i][4].replace(' ',''), progress, master)
                except:

                    error_message=str(i - 1)+" - "+data[i][0]+" "+data[i][2]+"의 공제 불공제 변경에 실패했습니다 - "+data[i][4].replace(' ','')
                    progress=((i-2)/(len(data)-2))*100
                    self.display_progress(error_message, progress, master, True)
            else:
                error_message=str(i - 1)+" - "+data[i][0]+" "+data[i][2]+"의 값이 홈택스와 다릅니다 - "+data[i][4].replace(' ','')
                progress=((i-2)/(len(data)-2))*100
                self.display_progress(error_message, progress, master, True)
                    
            #마지막 항목 체크  
            if i==len(data)-1:
                element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[5]/div/span/input')
                self.driver.execute_script("arguments[0].click();", element)
                sleep(0.5)
                self.alert_check()
                self.display_progress("변경이 완료되었습니다.", 100, master)

                
                for j in range(2, len(data)):
                    if data[j][4].replace(' ','')=='공제':
                        total_deduction += int(data[j][3])

                self.display_progress("총 공제금액 : "+"{:,}".format(total_deduction), 100, master)

                if ( master.total_deduction != 0):
                    self.display_progress("공제금액 합계 : "+ "{:,}".format(total_deduction + master.total_deduction), 100, master)

            else:
                #페이지 넘기기
                if i%20==1:
                    element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[5]/div/span/input')
                    self.driver.execute_script("arguments[0].click();", element)
                    sleep(0.5)
                    self.alert_check()
                    count_page+=1
                    if count_page==11:
                        element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[3]/div/div/div[1]/ul/li[13]/a')
                        self.driver.execute_script("arguments[0].click();", element)
                        count_page-=10
                    else :
                        element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[3]/div/div/div[1]/ul/li['+str(count_page+2)+']/a')
                        self.driver.execute_script("arguments[0].click();", element)
                    sleep(0.5)
                    self.alert_check()
                    while True:
                        try:
                            element = self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div/div[2]/div[4]/div[2]/div/div[1]/div/table/thead/tr/th[1]/input')
                            self.driver.execute_script("arguments[0].click();", element)
                            break
                        except Exception as e:
                            pass
        return total_deduction
                    
    # 진행상황 표시
    def display_progress(self, progress_text, progress, master, isFail=False):
        master.listbox2.insert(0, progress_text)
        if isFail:
            master.listbox2.itemconfig(0, {'fg': 'red'})
        master.progress_var.set(progress)
        master.progress_label.config(text=f"{progress:.1f}%")
        master.update_idletasks()

    def alert_check(self):
        # 경고창 확인
        try:
            da = Alert(self.driver)
            da.accept()
        except:
            pass

    def open_chrome(self):
        # 크롬 드라이버 오픈
        options = Options()
        chrome_args = ["--disable-logging","--log-level=3","--no-sandbox","--disable-gpu","--disable-extensions","--remote-debugging-port=0"]
        for arg in chrome_args:
            options.add_argument(arg)
        options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])  # 자동화 표시 제거
        options.add_experimental_option("useAutomationExtension", False)  # 자동화 확장 비활성화
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(options=options)
        driver.get(url="https://www.hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml")
        return driver

    def data_load_from_excel(self, filename):
        # 엑셀에서 데이터 로드
        try:
            workbook = xlrd.open_workbook(filename)
        except:
            return
        
        worksheet = workbook.sheet_by_index(0)
        all_values=[]
        for row in range(worksheet.nrows):
            row_value = []
            for col in range(worksheet.ncols):
                row_value.append(worksheet.cell(row, col).value)
            row_input=[]
            for i in range(len(row_value)):
                if i==0 or i==3 or i==4 or i==8 or i==12:
                    row_input.append(row_value[i])
            all_values.append(row_input)
        return all_values

    def ID_load_from_excel(self, filename):
        # 엑셀에서 데이터 로드
        id_wb = load_workbook(filename, data_only=True)
        id_ws = id_wb[id_wb.sheetnames[0]]
        all_values={}
        for row in id_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(str(cell.value))
            if row_value[0]!="None" and row_value[4]!="None":
                all_values[row_value[0]]=[row_value[1], row_value[2], row_value[3], row_value[4]]
        return all_values

    def login(self, selected_company):
        # 로그인
        id=self.ID_password[selected_company][1]
        key=self.ID_password[selected_company][2]
        res_no_front=self.ID_password[selected_company][3][0:6]
        res_no_back_first=self.ID_password[selected_company][3][6]
        
        # ID 입력
        self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[1]/div/div[1]/div[2]/div[2]/div[3]/div/div[1]/div[1]/ul/li[1]/div/input').send_keys(id)
        # Password 입력
        self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[1]/div/div[1]/div[2]/div[2]/div[3]/div/div[1]/div[1]/ul/li[2]/div/input').send_keys(key)
        self.driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div/div[1]/div/div[1]/div[2]/div[2]/div[3]/div/div[1]/div[2]/a').click()

        # 주민번호 앞자리, 뒷자리 첫번째 숫자 입력
        while True:
            try:
                self.driver.find_element(By.XPATH,'/html/body/div[6]/div[2]/div[1]/div/div/div[1]/div[2]/div/div[2]/div/ul/li[2]/div/input[1]').send_keys(res_no_front)
                self.driver.find_element(By.XPATH,'/html/body/div[6]/div[2]/div[1]/div/div/div[1]/div[2]/div/div[2]/div/ul/li[2]/div/input[2]').send_keys(res_no_back_first)
                self.driver.find_element(By.XPATH,'/html/body/div[6]/div[2]/div[1]/div/div/div[1]/div[2]/div/div[2]/div/div/input').click()
                break
            except Exception as e:
                pass 
        self.driver.switch_to.default_content()

    def close_chrome(self):
        # 크롬 드라이버 종료
        self.driver.quit()
    
    def change_site_url(self, url):
        try:
            if self.driver.current_url != url:
                self.driver.get(url)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"사이트 URL 변경 중 오류 발생: {e}")



#화면 전환
class display_interface(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title('공제 불공제 변경 프로그램')
        self.minsize(400, 300)  # 최소 사이즈
        self.resizable(False, False)
        self._frame = None
        self.ID_password = None
        self.selected_filename = None
        self.program_instance = hometax_program()
        self.switch_frame(hometax_program_GUI_first_window)

    def switch_frame(self, frame_class):
        try:
            new_frame = frame_class(self)
            if hasattr(self, "_current_frame") and self._current_frame is not None:
                self._current_frame.grid_forget()  # 현재 프레임을 숨김
            self._current_frame = new_frame
            self._current_frame.grid(row=0, column=0, sticky="nsew")
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")

#로그인 창
class hometax_program_GUI_first_window(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master = master
        self.program_instance = master.program_instance

        '''1. 프레임 생성'''
        # 상단 프레임 (LabelFrame)
        self.frm1 = tk.LabelFrame(self, text="로그인", pady=15, padx=15)
        self.frm1.grid(row=0, column=0, pady=10, padx=10, sticky="nswe")

        # 하단 프레임 (Frame)
        self.frm2 = tk.Frame(self, pady=10)
        self.frm2.grid(row=1, column=0, pady=10)

        '''2. 요소 생성'''
        # 레이블
        self.lbl1 = tk.Label(self.frm1, text='회사의 ID, password가 있는 엑셀 파일을 선택해 주세요.')
        self.lbl2 = tk.Label(self.frm1, text='로그인 하고자 하는 회사를 선택해 주세요.')

        # 리스트박스
        self.listbox1 = tk.Listbox(self.frm1, width=40, height=1)
        self.listbox2 = tk.Listbox(self.frm1, width=40)
        self.listbox2.bind("<Double-Button-1>", self.login)

        # 상단 버튼
        self.already_selected()
        self.select_button = tk.Button(self.frm1, text="찾아보기", width=12, command=self.select_file) 

        # 하단 버튼
        self.refresh_button = tk.Button(self.frm2, text="초기화", width=8, command=self.refresh)
        self.next_button = tk.Button(self.frm2, text="다음", width=8, command=self.switch_frame)
        self.quit_button = tk.Button(self.frm2, text="종료", width=8, command=self.all_quit)

        # 스크롤바 - 기능 연결
        self.scrollbar = tk.Scrollbar(self.frm1)
        self.scrollbar.config(command=self.listbox2.yview)
        self.listbox2.config(yscrollcommand=self.scrollbar.set)

        '''3. 요소 배치'''
        # 상단 프레임
        self.lbl1.grid(row=0, column=1, columnspan=2, sticky="w")
        self.listbox1.grid(row=1, column=1, columnspan=2, sticky="wens")
        self.lbl2.grid(row=2, column=1, columnspan=2, sticky="w")
        self.listbox2.grid(row=3, column=1, rowspan=2, sticky="wens")
        self.scrollbar.grid(row=3, column=2, rowspan=2, sticky="wens")
        self.select_button.grid(row=1, column=3)

        # 하단 프레임
        self.refresh_button.grid(row=0, column=3, sticky="wes")
        self.next_button.grid(row=0, column=6, sticky="ws")
        self.quit_button.grid(row=0, column=0, sticky="es")
        '''실행'''

    # ID 파일이 이미 선택되어 있을 경우
    def already_selected(self):
        try:
            if self.master.ID_password:
                self.listbox1.delete(0, "end")
                self.listbox1.insert(0, self.master.ID_password)
                self.display_company_list(self.master.ID_password)
            else :
                pass
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
        
    # 파일 선택
    def select_file(self):
        try:
            filename = askopenfilename(initialdir="./", filetypes=(("Excel files", ".xlsx .xls"), ('All files', '*.*')))
            if filename:
                self.listbox1.delete(0, "end")
                self.listbox1.insert(0, filename)
                self.display_company_list(filename)
                self.master.ID_password = filename
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    # 회사 표시
    def display_company_list(self, filename):
        try:
            company_list=self.program_instance.ID_load_from_excel(filename)
            self.program_instance.ID_password=company_list
            count=0
            for i in company_list.keys():
                display_content = i + " : " + company_list[i][0]
                self.listbox2.insert(count,display_content)
                count+=1
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    # 초기화
    def refresh(self):
        try:
            reply = messagebox.askyesno("초기화", "정말로 초기화 하시겠습니까?")
            if reply:
                self.listbox1.delete(0, "end")
                self.listbox2.delete(0, "end")
                messagebox.showinfo("Success", "초기화 되었습니다.")
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    # 로그인
    def login(self, event):
        try:
            selected_company = self.listbox2.get(self.listbox2.curselection())
            self.program_instance.login(selected_company.split(" : ")[0])
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    # 다음
    def switch_frame(self):
        try:
            self.master.switch_frame(hometax_program_GUI_second_window)
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")

    # 종료
    def all_quit(self):
        try:
            reply = messagebox.askyesno("종료", "정말로 종료 하시겠습니까?")
            if reply:
                self.program_instance.close_chrome()
                self.master.destroy()
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")

#공제 창
class hometax_program_GUI_second_window(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master = master
        self.program_instance = master.program_instance
        self.filename = self.master.selected_filename
        self.total_deduction = 0

        '''1. 프레임 생성'''
        # 상단 프레임 (LabelFrame)
        self.frm1 = tk.LabelFrame(self, text="공제 불공제 변경", pady=15, padx=15)   # pad 내부
        self.frm1.grid(row=0, column=0, pady=10, padx=10, sticky="nswe") # pad 내부

        # 하단 프레임 (Frame)
        self.frm2 = tk.Frame(self, pady=10)
        self.frm2.grid(row=1, column=0, pady=10)

        '''2. 요소 생성'''
        # 레이블
        self.lbl1 = tk.Label(self.frm1, text='자료가 들어있는 엑셀 파일을 선택해 주세요.')
        self.lbl2 = tk.Label(self.frm1, text='진행상황.')

        # 리스트박스
        self.listbox1 = tk.Listbox(self.frm1, width=40, height=1)
        self.listbox2 = tk.Listbox(self.frm1, width=40)

        # 진행상황 표시
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.frm1, variable=self.progress_var, maximum=100, length=200, mode="determinate")
        self.progress_label = tk.Label(self.frm1, text="0%")

        # 상단 버튼
        self.select_button = tk.Button(self.frm1, text="찾아보기", width=12, command=self.select_file) 

        # 실행, 사이트 버튼들 위한 서브 프레임
        self.button_frame = tk.Frame(self.frm1)
        self.button_frame.grid(row=3, column=3, rowspan=3, sticky="n", pady=5)

        # 버튼 폭 통일 (width=15 정도 추천)
        self.start_button = tk.Button(self.button_frame, text="변경 실행", width=12, command=self.change_deduction)
        self.site_button1 = tk.Button(self.button_frame, text="변경 사이트", width=12, command=lambda: self.program_instance.change_site_url("https://hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml&tmIdx=46&tm2lIdx=4608020000&tm3lIdx=4608020100"))
        self.site_button2 = tk.Button(self.button_frame, text="조회 사이트", width=12, command=lambda: self.program_instance.change_site_url("https://hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml&tmIdx=46&tm2lIdx=4608020000&tm3lIdx=4608020200"))

        # 버튼 배치 (패킹)
        self.start_button.pack(pady=2)
        self.site_button1.pack(pady=2)
        self.site_button2.pack(pady=2)

        # 하단 버튼
        self.refresh_button = tk.Button(self.frm2, text="초기화", width=8, command=self.refresh)
        self.next_button = tk.Button(self.frm2, text="이전", width=8, command=self.switch_frame)
        self.quit_button = tk.Button(self.frm2, text="종료", width=8, command=self.all_quit)

        # 스크롤바 - 기능 연결
        self.scrollbar = tk.Scrollbar(self.frm1, orient="vertical")
        self.scrollbar.config(command=self.listbox2.yview)

        self.scrollbar_x = tk.Scrollbar(self.frm1, orient="horizontal")
        self.scrollbar_x.config(command=self.listbox2.xview)

        self.listbox2.config(yscrollcommand=self.scrollbar.set)
        self.listbox1.config(xscrollcommand=self.scrollbar_x.set)

        '''3. 요소 배치'''
        # 상단 프레임
        self.lbl1.grid(row=0, column=1, columnspan=2, sticky="w")
        self.lbl2.grid(row=2, column=1, columnspan=2, sticky="w")
        self.listbox1.grid(row=1, column=1, columnspan=2, sticky="wens")
        self.listbox2.grid(row=3, column=1, rowspan=2, sticky="wens")
        self.scrollbar.grid(row=3, column=2, rowspan=2, sticky="ns")
        self.scrollbar_x.grid(row=5, column=1, sticky="ew")
        self.progress_bar.grid(row=6, column=1, columnspan=2, sticky="wens")
        self.progress_label.grid(row=7, column=1, columnspan=2, sticky="w")
        self.select_button.grid(row=1, column=3)

        # 하단 프레임
        self.refresh_button.grid(row=0, column=3, sticky="wes")
        self.next_button.grid(row=0, column=6, sticky="ws")
        self.quit_button.grid(row=0, column=0, sticky="es")
        '''실행'''

    #파일 선택
    def select_file(self):
        try:
            if self.filename is None:
                self.filename = askopenfilename(initialdir="./", filetypes=(("Excel files", ".xlsx .xls"), ('All files', '*.*')))
            else:
                # 파일이 이미 선택되어 있을 경우, 그 디렉토리에서 파일 선택
                self.filename = askopenfilename(initialdir=os.path.dirname(self.filename), filetypes=(("Excel files", ".xlsx .xls"), ('All files', '*.*')))
            if self.filename:
                self.listbox1.delete(0, "end")
                self.listbox1.insert(0, self.filename)
                self.master.selected_filename = self.filename
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
            
    #공제 불공제 변경
    def change_deduction(self):
        try:    
            data_from_excel=self.program_instance.data_load_from_excel(self.listbox1.get(0))
            self.total_deduction += self.program_instance.change_deduction(data_from_excel, self)
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")

    #초기화    
    def refresh(self):
        try:
            reply = messagebox.askyesno("초기화", "정말로 초기화 하시겠습니까?")
            if reply:
                self.listbox1.delete(0, "end")
                self.listbox2.delete(0, "end")
                messagebox.showinfo("Success", "초기화 되었습니다.")
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    #이전
    def switch_frame(self):
        try:
            self.master.switch_frame(hometax_program_GUI_first_window)
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")
    #종료   
    def all_quit(self):
        try:
            reply = messagebox.askyesno("종료", "정말로 종료 하시겠습니까?")
            if reply:
                self.program_instance.close_chrome()
                self.master.destroy()   
        except:
            messagebox.showinfo("Error", "오류가 발생했습니다.")

if __name__ == "__main__":
    os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
    warnings.filterwarnings('ignore')
    app = display_interface()
    app.mainloop()